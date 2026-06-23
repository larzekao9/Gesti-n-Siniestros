"""Report domain service — CU-22 operational reports (PDF / Excel)."""

import io
from datetime import date, datetime, timezone
from uuid import UUID

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.claim import Claim
from app.models.policyholder import Policyholder
from app.services.analytics_service import analytics_service
from app.services.audit_service import AuditService
from app.services.claim_service import ClaimService

audit_service = AuditService()
claim_service = ClaimService()

_STATUS_LABELS = {
    "registered": "Registrado",
    "in_review": "En revisión",
    "observed": "Observado",
    "docs_pending": "Docs. pendientes",
    "in_evaluation": "En evaluación",
    "approved": "Aprobado",
    "rejected": "Rechazado",
    "closed": "Cerrado",
}

_COLUMNS = [
    ("claim_number", "Expediente"),
    ("status", "Estado"),
    ("policyholder", "Asegurado"),
    ("accident_date", "Fecha siniestro"),
    ("accident_location", "Lugar"),
    ("created_at", "Registrado"),
]


class ReportService:
    """Builds operational claim reports in PDF or Excel.

    Reúsa las queries de ``AnalyticsService`` (KPIs de cabecera) y
    ``ClaimService.search`` (detalle de filas). Genera de forma síncrona;
    para volúmenes grandes existe un Celery task wrapper en
    ``app/tasks/report_generation.py``.
    """

    async def _collect(
        self,
        db: AsyncSession,
        *,
        tenant_id: UUID,
        from_date: date | None,
        to_date: date | None,
        status: str | None,
        analyst_id: str | None,
        q: str | None = None,
        policyholder_id: str | None = None,
        supervisor_id: str | None = None,
    ) -> tuple[dict, list[dict]]:
        kpis = await analytics_service.get_kpis(
            db,
            tenant_id=tenant_id,
            from_date=from_date,
            to_date=to_date,
            analyst_id=analyst_id,
        )

        # Traemos todas las filas que matchean (paginando alto; académico).
        rows: list[Claim] = []
        page = 1
        while True:
            batch, total = await claim_service.search(
                db,
                tenant_id=tenant_id,
                q=q,
                status=status,
                from_date=from_date,
                to_date=to_date,
                analyst_id=analyst_id,
                supervisor_id=supervisor_id,
                policyholder_id=policyholder_id,
                page=page,
                limit=200,
            )
            rows.extend(batch)
            if len(rows) >= total or not batch:
                break
            page += 1

        # Resolvemos nombres de asegurados en un solo query.
        ph_ids = {c.policyholder_id for c in rows}
        names: dict[UUID, str] = {}
        if ph_ids:
            res = await db.execute(
                select(Policyholder.id, Policyholder.full_name).where(
                    Policyholder.id.in_(ph_ids), Policyholder.tenant_id == tenant_id
                )
            )
            names = {pid: pname for pid, pname in res.all()}

        table_rows: list[dict] = []
        for c in rows:
            status_val = c.status.value if hasattr(c.status, "value") else str(c.status)
            table_rows.append(
                {
                    "claim_number": c.claim_number,
                    "status": _STATUS_LABELS.get(status_val, status_val),
                    "policyholder": names.get(c.policyholder_id, "—"),
                    "accident_date": c.accident_date.isoformat() if c.accident_date else "",
                    "accident_location": c.accident_location or "",
                    "created_at": c.created_at.date().isoformat() if c.created_at else "",
                }
            )
        return kpis, table_rows

    def _render_pdf(self, kpis: dict, rows: list[dict], title: str) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
            leftMargin=1.5 * cm,
            rightMargin=1.5 * cm,
        )
        styles = getSampleStyleSheet()
        elements: list = []

        elements.append(Paragraph(title, styles["Title"]))
        generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        elements.append(Paragraph(f"Generado: {generated}", styles["Normal"]))
        elements.append(Spacer(1, 0.4 * cm))

        kpi_line = (
            f"Total expedientes: {kpis['total_claims']} &nbsp;|&nbsp; "
            f"Abiertos: {kpis['open_claims']} &nbsp;|&nbsp; "
            f"Tasa de aprobación: {kpis['approval_rate'] * 100:.1f}% &nbsp;|&nbsp; "
            f"Rechazo en intake: {kpis['intake_rejection_rate'] * 100:.1f}% &nbsp;|&nbsp; "
            f"Días prom. a decisión: {kpis['avg_days_to_decision'] if kpis['avg_days_to_decision'] is not None else '—'}"
        )
        elements.append(Paragraph(kpi_line, styles["Normal"]))
        elements.append(Spacer(1, 0.5 * cm))

        header = [label for _, label in _COLUMNS]
        data = [header]
        if rows:
            for r in rows:
                data.append([r[key] for key, _ in _COLUMNS])
        else:
            data.append(["Sin datos para el rango seleccionado"] + [""] * (len(header) - 1))

        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1d4ed8")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        elements.append(table)
        doc.build(elements)
        return buffer.getvalue()

    def _render_xlsx(self, kpis: dict, rows: list[dict], title: str) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()

        ws_kpis = wb.active
        ws_kpis.title = "Resumen"
        ws_kpis["A1"] = title
        ws_kpis["A1"].font = Font(bold=True, size=14)
        ws_kpis["A3"] = "Generado"
        ws_kpis["B3"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        kpi_pairs = [
            ("Total expedientes", kpis["total_claims"]),
            ("Total solicitudes", kpis["total_requests"]),
            ("Expedientes abiertos", kpis["open_claims"]),
            ("Tasa de aprobación", f"{kpis['approval_rate'] * 100:.1f}%"),
            ("Rechazo en intake", f"{kpis['intake_rejection_rate'] * 100:.1f}%"),
            (
                "Días promedio a decisión",
                kpis["avg_days_to_decision"] if kpis["avg_days_to_decision"] is not None else "—",
            ),
        ]
        for i, (label, value) in enumerate(kpi_pairs, start=5):
            ws_kpis[f"A{i}"] = label
            ws_kpis[f"A{i}"].font = Font(bold=True)
            ws_kpis[f"B{i}"] = value
        ws_kpis.column_dimensions["A"].width = 26
        ws_kpis.column_dimensions["B"].width = 26

        ws = wb.create_sheet("Expedientes")
        header_fill = PatternFill("solid", fgColor="1D4ED8")
        header_font = Font(bold=True, color="FFFFFF")
        for col, (_, label) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.fill = header_fill
            cell.font = header_font
        for r_idx, r in enumerate(rows, start=2):
            for c_idx, (key, _) in enumerate(_COLUMNS, start=1):
                ws.cell(row=r_idx, column=c_idx, value=r[key])
        for col, (key, label) in enumerate(_COLUMNS, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max(
                14, len(label) + 4
            )

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    async def build_claims_report(
        self,
        db: AsyncSession,
        *,
        tenant_id: UUID,
        actor_user_id: UUID,
        report_format: str,
        from_date: date | None = None,
        to_date: date | None = None,
        status: str | None = None,
        analyst_id: str | None = None,
        q: str | None = None,
        policyholder_id: str | None = None,
        supervisor_id: str | None = None,
    ) -> tuple[bytes, str, str]:
        """Return ``(content, media_type, filename)`` for the requested report."""
        kpis, rows = await self._collect(
            db,
            tenant_id=tenant_id,
            from_date=from_date,
            to_date=to_date,
            status=status,
            analyst_id=analyst_id,
            q=q,
            policyholder_id=policyholder_id,
            supervisor_id=supervisor_id,
        )
        title = "Reporte operativo de siniestros"
        stamp = datetime.now(timezone.utc).strftime("%Y%m%d")

        if report_format == "xlsx":
            content = self._render_xlsx(kpis, rows, title)
            media_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename = f"reporte_siniestros_{stamp}.xlsx"
        else:
            content = self._render_pdf(kpis, rows, title)
            media_type = "application/pdf"
            filename = f"reporte_siniestros_{stamp}.pdf"

        await audit_service.write(
            db,
            tenant_id=tenant_id,
            actor_user_id=actor_user_id,
            action="REPORT_GENERATE",
            entity_type="report",
            payload_diff={
                "format": report_format,
                "rows": len(rows),
                "filters": {
                    "from": from_date.isoformat() if from_date else None,
                    "to": to_date.isoformat() if to_date else None,
                    "status": status,
                    "analyst": analyst_id,
                    "supervisor": supervisor_id,
                    "policyholder": policyholder_id,
                    "q": q,
                },
            },
        )
        return content, media_type, filename


report_service = ReportService()
